"""Excel import service — mirrors ExcelImportService.java using openpyxl."""

import openpyxl
from io import BytesIO
from sqlalchemy.orm import Session

from services import employee_service


def import_employees(db: Session, file_bytes: bytes, actor: str) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = 0
    failed = 0
    errors = []

    headers = [cell.value for cell in ws[1]]
    header_map = {h.strip().lower(): i for i, h in enumerate(headers) if h}

    def get_val(row, key):
        idx = header_map.get(key.lower())
        if idx is not None and idx < len(row):
            v = row[idx]
            return str(v).strip() if v is not None else None
        return None

    for row_idx, row in enumerate(rows, start=2):
        try:
            dto = {
                "name": get_val(row, "name"),
                "companyEmail": get_val(row, "companyemail") or get_val(row, "company_email"),
                "personalEmail": get_val(row, "personalemail") or get_val(row, "personal_email"),
                "phoneNumber": get_val(row, "phonenumber") or get_val(row, "phone_number"),
                "address": get_val(row, "address"),
                "department": get_val(row, "department"),
                "designation": get_val(row, "designation"),
                "skills": get_val(row, "skills"),
                "gender": get_val(row, "gender"),
                "description": get_val(row, "description"),
                "roles": ["EMPLOYEE"],
            }
            doj = get_val(row, "dateofjoin") or get_val(row, "date_of_join")
            if doj:
                from datetime import date as dt_date
                dto["dateOfJoin"] = doj if isinstance(doj, dt_date) else doj

            if not dto.get("name") or not dto.get("companyEmail"):
                errors.append(f"Row {row_idx}: name and companyEmail are required")
                failed += 1
                continue

            employee_service.create_employee(db, dto, actor)
            created += 1
        except Exception as e:
            failed += 1
            errors.append(f"Row {row_idx}: {str(e)}")

    wb.close()
    return {"created": created, "failed": failed, "errors": errors}
